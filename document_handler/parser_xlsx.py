import openpyxl
import os
from typing import Dict, Any, Optional, List, Union
from utils.logger import get_logger

logger = get_logger(__name__)

class XlsxParser:
    """Parser for Excel XLSX/XLS files"""
    
    def __init__(self):
        self.max_file_size = 50 * 1024 * 1024  # 50MB
        self.max_rows_per_sheet = 1000  # Limit rows to prevent memory issues
        
    def parse_xlsx(self, file_path: str) -> Optional[Dict[str, Any]]:
        """Parse XLSX file and extract data and metadata"""
        try:
            if not os.path.exists(file_path):
                logger.error(f"Excel file not found: {file_path}")
                return None
            
            # Check file size
            file_size = os.path.getsize(file_path)
            if file_size > self.max_file_size:
                logger.error(f"Excel file too large: {file_size} bytes")
                return None
            
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            result = {
                'type': 'xlsx',
                'extracted_text': '',
                'metadata': {
                    'size': file_size,
                    'file_path': file_path,
                    'sheets': {},
                    'total_sheets': len(workbook.sheetnames)
                },
                'sheets': {},
                'summary': {},
                'success': True
            }
            
            # Extract data from each sheet
            all_text = ""
            total_rows = 0
            total_cells = 0
            
            for sheet_name in workbook.sheetnames:
                sheet_data = self._extract_sheet_data(workbook[sheet_name], sheet_name)
                result['sheets'][sheet_name] = sheet_data
                
                # Update metadata
                result['metadata']['sheets'][sheet_name] = {
                    'rows': sheet_data['row_count'],
                    'columns': sheet_data['column_count'],
                    'data_rows': len(sheet_data['data'])
                }
                
                total_rows += sheet_data['row_count']
                total_cells += sheet_data['cell_count']
                
                # Add sheet data to extracted text
                all_text += f"\n--- Sheet: {sheet_name} ---\n"
                for row in sheet_data['data'][:50]:  # Limit to first 50 rows for text
                    row_text = " | ".join(str(cell) for cell in row if cell is not None)
                    if row_text.strip():
                        all_text += row_text + "\n"
            
            result['extracted_text'] = all_text.strip()
            result['metadata']['total_rows'] = total_rows
            result['metadata']['total_cells'] = total_cells
            
            # Create summary
            result['summary'] = self._create_workbook_summary(result['sheets'])
            
            # Analyze content
            content_analysis = self._analyze_excel_content(result['sheets'])
            result['content_analysis'] = content_analysis
            
            workbook.close()
            
            logger.info(f"Successfully parsed Excel: {file_path}, {result['metadata']['total_sheets']} sheets, {total_rows} rows")
            
            return result
            
        except Exception as e:
            logger.error(f"Error parsing Excel {file_path}: {e}")
            return {
                'type': 'xlsx',
                'extracted_text': '',
                'metadata': {'error': str(e)},
                'success': False
            }
    
    def extract_numerical_data(self, file_path: str) -> Optional[Dict[str, List[Dict[str, Any]]]]:
        """Extract numerical data from Excel file"""
        try:
            result = self.parse_xlsx(file_path)
            if not result or not result['success']:
                return None
            
            numerical_data = {}
            
            for sheet_name, sheet_data in result['sheets'].items():
                numerical_data[sheet_name] = []
                
                for row_idx, row in enumerate(sheet_data['data']):
                    for col_idx, cell in enumerate(row):
                        if isinstance(cell, (int, float)) and cell != 0:
                            numerical_data[sheet_name].append({
                                'row': row_idx + 1,
                                'column': col_idx + 1,
                                'value': cell,
                                'formatted': self._format_number(cell)
                            })
            
            return numerical_data
            
        except Exception as e:
            logger.error(f"Error extracting numerical data: {e}")
            return None
    
    def find_financial_tables(self, file_path: str) -> Optional[List[Dict[str, Any]]]:
        """Find tables that look like financial/cost data"""
        try:
            result = self.parse_xlsx(file_path)
            if not result or not result['success']:
                return None
            
            financial_tables = []
            
            for sheet_name, sheet_data in result['sheets'].items():
                # Look for patterns that indicate financial data
                financial_indicators = ['total', 'subtotal', 'cost', 'price', 'amount', 'biaya', 'harga', 'jumlah', 'rp']
                
                data = sheet_data['data']
                for row_idx, row in enumerate(data):
                    row_text = ' '.join(str(cell).lower() for cell in row if cell is not None)
                    
                    # Check if row contains financial indicators
                    if any(indicator in row_text for indicator in financial_indicators):
                        # Extract surrounding context (previous and next few rows)
                        start_row = max(0, row_idx - 2)
                        end_row = min(len(data), row_idx + 10)
                        
                        table_data = data[start_row:end_row]
                        
                        financial_tables.append({
                            'sheet': sheet_name,
                            'start_row': start_row + 1,
                            'end_row': end_row,
                            'header_row': row_idx + 1,
                            'data': table_data,
                            'indicators_found': [ind for ind in financial_indicators if ind in row_text]
                        })
                        
                        break  # Only find first financial table per sheet
            
            return financial_tables
            
        except Exception as e:
            logger.error(f"Error finding financial tables: {e}")
            return None
    
    def _extract_sheet_data(self, worksheet, sheet_name: str) -> Dict[str, Any]:
        """Extract data from a single worksheet"""
        try:
            # Get dimensions
            max_row = min(worksheet.max_row or 0, self.max_rows_per_sheet)
            max_col = worksheet.max_column or 0
            
            data = []
            cell_count = 0
            
            # Extract cell values
            for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True):
                row_data = []
                for cell_value in row:
                    if cell_value is not None:
                        cell_count += 1
                        
                        # Handle different data types
                        if isinstance(cell_value, (int, float)):
                            row_data.append(cell_value)
                        else:
                            row_data.append(str(cell_value).strip())
                    else:
                        row_data.append(None)
                
                # Only add rows that have some data
                if any(cell is not None for cell in row_data):
                    data.append(row_data)
            
            # Analyze data types in the sheet
            data_analysis = self._analyze_sheet_data_types(data)
            
            return {
                'name': sheet_name,
                'data': data,
                'row_count': max_row,
                'column_count': max_col,
                'cell_count': cell_count,
                'data_analysis': data_analysis
            }
            
        except Exception as e:
            logger.error(f"Error extracting sheet data for {sheet_name}: {e}")
            return {
                'name': sheet_name,
                'data': [],
                'row_count': 0,
                'column_count': 0,
                'cell_count': 0,
                'error': str(e)
            }
    
    def _analyze_sheet_data_types(self, data: List[List[Any]]) -> Dict[str, Any]:
        """Analyze data types and patterns in sheet data"""
        analysis = {
            'total_cells_with_data': 0,
            'numeric_cells': 0,
            'text_cells': 0,
            'date_cells': 0,
            'formula_cells': 0,
            'monetary_values': [],
            'large_numbers': [],
            'percentages': []
        }
        
        import re
        from datetime import datetime
        
        for row in data:
            for cell in row:
                if cell is None:
                    continue
                
                analysis['total_cells_with_data'] += 1
                
                if isinstance(cell, (int, float)):
                    analysis['numeric_cells'] += 1
                    
                    # Check for monetary values (large numbers)
                    if cell > 100000:  # Assuming values > 100k might be monetary
                        analysis['large_numbers'].append(float(cell))
                    
                    # Check for percentages (values between 0 and 1, or 0 and 100)
                    if 0 <= cell <= 1:
                        analysis['percentages'].append(float(cell))
                    elif 0 <= cell <= 100 and cell != int(cell):
                        analysis['percentages'].append(float(cell))
                
                elif isinstance(cell, datetime):
                    analysis['date_cells'] += 1
                
                elif isinstance(cell, str):
                    analysis['text_cells'] += 1
                    
                    # Look for monetary indicators in text
                    if re.search(r'(?:rp|rupiah|\$|usd)', cell.lower()):
                        # Try to extract numbers from the text
                        numbers = re.findall(r'\d+[.,]?\d*', cell)
                        for num_str in numbers:
                            try:
                                num = float(num_str.replace(',', ''))
                                if num > 1000:
                                    analysis['monetary_values'].append(num)
                            except ValueError:
                                pass
                    
                    # Check for percentage indicators
                    if '%' in cell:
                        numbers = re.findall(r'\d+[.,]?\d*', cell)
                        for num_str in numbers:
                            try:
                                num = float(num_str.replace(',', ''))
                                analysis['percentages'].append(num)
                            except ValueError:
                                pass
        
        # Limit lists to prevent memory issues
        analysis['monetary_values'] = analysis['monetary_values'][:50]
        analysis['large_numbers'] = analysis['large_numbers'][:50]
        analysis['percentages'] = analysis['percentages'][:20]
        
        return analysis
    
    def _create_workbook_summary(self, sheets: Dict[str, Any]) -> Dict[str, Any]:
        """Create a summary of the workbook content"""
        summary = {
            'total_sheets': len(sheets),
            'total_data_rows': 0,
            'total_numeric_values': 0,
            'largest_values': [],
            'sheet_purposes': {}
        }
        
        all_large_numbers = []
        
        for sheet_name, sheet_data in sheets.items():
            if 'data_analysis' not in sheet_data:
                continue
            
            analysis = sheet_data['data_analysis']
            summary['total_data_rows'] += len(sheet_data['data'])
            summary['total_numeric_values'] += analysis['numeric_cells']
            
            # Collect large numbers
            all_large_numbers.extend(analysis.get('large_numbers', []))
            all_large_numbers.extend(analysis.get('monetary_values', []))
            
            # Guess sheet purpose based on content
            purpose = self._guess_sheet_purpose(sheet_name, sheet_data['data'])
            summary['sheet_purposes'][sheet_name] = purpose
        
        # Sort and limit large numbers
        all_large_numbers.sort(reverse=True)
        summary['largest_values'] = all_large_numbers[:20]
        
        return summary
    
    def _analyze_excel_content(self, sheets: Dict[str, Any]) -> Dict[str, Any]:
        """Analyze Excel content to determine document type"""
        analysis = {
            'document_type': 'unknown',
            'confidence': 0.0,
            'key_indicators': [],
            'likely_purpose': 'data_storage'
        }
        
        # Document type indicators
        type_indicators = {
            'budget': ['budget', 'anggaran', 'rab', 'cost', 'biaya', 'total', 'subtotal'],
            'inventory': ['stock', 'inventory', 'item', 'quantity', 'qty', 'material'],
            'financial': ['income', 'expense', 'profit', 'loss', 'revenue', 'pendapatan'],
            'schedule': ['schedule', 'timeline', 'jadwal', 'date', 'week', 'month'],
            'analysis': ['analysis', 'trend', 'comparison', 'variance', 'ratio'],
            'report': ['report', 'summary', 'dashboard', 'kpi', 'metric']
        }
        
        # Count indicators across all sheets
        indicator_counts = {doc_type: 0 for doc_type in type_indicators}
        
        for sheet_name, sheet_data in sheets.items():
            sheet_text = sheet_name.lower()
            
            # Check sheet name for indicators
            for doc_type, indicators in type_indicators.items():
                for indicator in indicators:
                    if indicator in sheet_text:
                        indicator_counts[doc_type] += 2  # Sheet name is strong indicator
            
            # Check data content for indicators
            for row in sheet_data.get('data', [])[:10]:  # Check first 10 rows
                for cell in row:
                    if cell and isinstance(cell, str):
                        cell_text = cell.lower()
                        for doc_type, indicators in type_indicators.items():
                            for indicator in indicators:
                                if indicator in cell_text:
                                    indicator_counts[doc_type] += 1
        
        # Determine most likely document type
        if indicator_counts:
            max_type = max(indicator_counts, key=indicator_counts.get)
            max_count = indicator_counts[max_type]
            
            if max_count > 0:
                analysis['document_type'] = max_type
                analysis['confidence'] = min(max_count / 10.0, 1.0)  # Normalize
                analysis['key_indicators'] = [ind for ind in type_indicators[max_type] 
                                            if any(ind in str(cell).lower() 
                                                 for sheet in sheets.values() 
                                                 for row in sheet.get('data', [])[:10] 
                                                 for cell in row if cell)]
        
        return analysis
    
    def _guess_sheet_purpose(self, sheet_name: str, data: List[List[Any]]) -> str:
        """Guess the purpose of a sheet based on name and content"""
        name_lower = sheet_name.lower()
        
        # Common sheet purposes
        purpose_patterns = {
            'summary': ['summary', 'ringkasan', 'total', 'overview'],
            'data': ['data', 'raw', 'input', 'master'],
            'calculation': ['calc', 'perhitungan', 'formula', 'analysis'],
            'chart': ['chart', 'graph', 'grafik', 'visualization'],
            'report': ['report', 'laporan', 'output', 'result'],
            'budget': ['budget', 'anggaran', 'cost', 'biaya'],
            'schedule': ['schedule', 'jadwal', 'timeline', 'planning']
        }
        
        for purpose, patterns in purpose_patterns.items():
            if any(pattern in name_lower for pattern in patterns):
                return purpose
        
        # Analyze first few rows for clues
        if data:
            first_rows_text = ' '.join(str(cell).lower() for row in data[:5] for cell in row if cell)
            
            for purpose, patterns in purpose_patterns.items():
                if any(pattern in first_rows_text for pattern in patterns):
                    return purpose
        
        return 'unknown'
    
    def _format_number(self, number: Union[int, float]) -> str:
        """Format number for display"""
        if isinstance(number, int):
            return f"{number:,}"
        elif isinstance(number, float):
            if number.is_integer():
                return f"{int(number):,}"
            else:
                return f"{number:,.2f}"
        else:
            return str(number)
